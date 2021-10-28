from typing import Optional

from openpyxl import load_workbook

import db
from models import GoodsItem
from io import BytesIO
import re


def excel_row_to_item(row: tuple) -> Optional[GoodsItem]:
    name = row[0].strip() if row[0] else None
    description = row[1].strip() if row[1] else None
    category = row[2]
    all_prices_str = " ".join(str(x) for x in filter(lambda x: x is not None, row[3:]))
    prices = list(map(float, re.findall(r'\d+', all_prices_str)))
    # prices = list(map(
    #     lambda p: float(p.replace("р", "").replace("Р", "").strip()),
    #     filter(lambda p: p is not None, row[3:])
    # ))
    # price = min(prices) if prices else None  # min price
    price = round(sum(prices)/len(prices), 2) if prices else None  # mean price
    if all(x is None for x in [name, description, price]):  # все поля пустые
        return None
    item = GoodsItem(
        name=name,
        description=description,
        category=category,
        price=price,
    )

    return item


def import_goods_from_xlx(goods_xlsx: BytesIO):
    wb = load_workbook(goods_xlsx, read_only=True)
    ws = wb.active

    print("Parsing ")
    for row in ws.iter_rows(min_row=2, min_col=16, max_col=23, values_only=True):
        item = excel_row_to_item(row)
        if item is None:
            continue
        # print(item.to_dict())
        db.add_goods_item(item)
        # print(item.to_dict())


# if __name__ == "__main__":
#     # for test purposes
#     test_file = "Kazan_v1.xlsx"
#     with open(test_file, 'rb') as f:
#         buf = BytesIO(f.read())
#         import_goods_from_xlx(buf)

